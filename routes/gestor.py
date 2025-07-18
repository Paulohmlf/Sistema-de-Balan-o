# routes/gestor.py

from flask import (
    Blueprint, render_template, request, redirect, url_for,
    session, flash, jsonify, make_response, send_file
)
from io import BytesIO
from routes.auth import login_required
import data_manager # Certifique-se de que data_manager.py foi atualizado com a coluna VALIDADO
from data_manager import find_item, update_item, get_all_users, update_user_password
from werkzeug.security import generate_password_hash
import pandas as pd
from datetime import datetime
import io
from werkzeug.utils import secure_filename
import openpyxl
from functools import wraps
import logging
from copy import copy
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

gestor_bp = Blueprint('gestor', __name__, url_prefix='/')

# Configurações para Excel
EXCEL_CONFIG = {
    'NOME_ABA_EXCEL': "INVENTÁRIO",
    'LINHA_CABECALHO_EXCEL': 15,
    'COL_EXCEL_CHAVE_CODIGO': "CÓDIGO",
    'COL_EXCEL_CHAVE_LOTE': "LOTE",
    'COL_EXCEL_DESTINO_RESP_CONTAGEM': "RESP. PELA CONTAGEM",
    'COL_EXCEL_DESTINO_INVENTARIO': "INVENTÁRIO",
    'COL_EXCEL_DESTINO_CONFERENCIA': "CONFERENCIA"
}

# Decorator para tratamento de erros
def handle_exceptions(redirect_route='gestor.index'):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            try:
                return f(*args, **kwargs)
            except ValueError as e:
                logger.warning(f"Erro de validação em {f.__name__}: {e}")
                flash("Dados inválidos fornecidos", "warning")
                return redirect(url_for(redirect_route))
            except Exception as e:
                logger.error(f"Erro inesperado em {f.__name__}: {e}")
                flash("Erro interno do sistema", "danger")
                return redirect(url_for(redirect_route))
        return decorated_function
    return decorator

def get_manufacturers_status():
    """Versão otimizada com uma única passada pelos dados"""
    fabricantes_status = {}

    # Uma única iteração sobre todos os dados
    for item in data_manager.INVENTORY_DATA:
        fab = str(item.get('FABRICANTE', '')).strip()
        if not fab:
            continue

        if fab not in fabricantes_status:
            fabricantes_status[fab] = {
                'status': 'Pendente',
                'user': '',
                'items_concluidos': 0,
                'items_total': 0,
                'has_recount': False,
                'has_in_progress': False,
                'items_validados_sim': 0, # Novo contador para itens validados como 'Sim'
                'total_itens_do_fabricante': 0 # Novo contador para total de itens do fabricante
            }

        status_info = fabricantes_status[fab]
        status_info['items_total'] += 1

        item_status = str(item.get('STATUS'))
        if item_status == 'Concluído':
            status_info['items_concluidos'] += 1
            if not status_info['user']:
                status_info['user'] = item.get('CONTADOR', '')
            # Verifica o status 'VALIDADO' para itens concluídos
            if str(item.get('VALIDADO', 'Não')) == 'Sim':
                status_info['items_validados_sim'] += 1
        elif item_status == 'Recontagem':
            status_info['has_recount'] = True
            status_info['user'] = item.get('CONTADOR_ORIGINAL', '')
        elif item_status == 'Em Andamento':
            status_info['has_in_progress'] = True
            status_info['user'] = item.get('CONTADOR', '')
        
        status_info['total_itens_do_fabricante'] += 1 # Conta todos os itens para cada fabricante

    # Determina status final baseado nas contagens
    for fab, info in fabricantes_status.items():
        if info['has_recount']:
            info['status'] = 'Em Andamento'
        elif info['has_in_progress']:
            info['status'] = 'Em Andamento'
        elif info['items_concluidos'] == info['items_total']:
            info['status'] = 'Concluído'
        
        # Determina se todos os itens estão validados como 'Sim'
        if info['items_concluidos'] == info['items_total'] and \
           info['items_validados_sim'] == info['items_total'] and \
           info['items_total'] > 0: # Garante que há itens para validar
            info['status_verificado'] = True
        else:
            info['status_verificado'] = False
            
    return fabricantes_status

def calculate_item_differences(inventario_filtrado):
    """Calcula diferenças numéricas para itens do inventário"""
    for item in inventario_filtrado:
        estoque_sistema = item.get('ESTOQUE_SISTEMA', 0)
        inventario_final = item.get('INVENTARIO', 0)
        diferenca_num = inventario_final - estoque_sistema
        item['DIFERENCA_NUMERICA_CALC'] = diferenca_num
        if diferenca_num > 0:
            item['STATUS_DIFERENCA_CALC'] = 'Mais'
        elif diferenca_num < 0:
            item['STATUS_DIFERENCA_CALC'] = 'Menos'
        else:
            item['STATUS_DIFERENCA_CALC'] = 'OK'

def mark_notifications_as_seen():
    """Marca notificações como vistas"""
    session['seen_notifications'] = len(data_manager.NOTIFICATIONS)

@gestor_bp.route('/', methods=['GET', 'POST'])
@login_required(role="gestor")
@handle_exceptions()
def index():
    mark_notifications_as_seen()
    fabricantes_status_dict = get_manufacturers_status()
    inventario_filtrado = []
    fabricantes_selecionados_for_template = []
    percentual_divergencia_str = request.form.get(
        'percentual_divergencia',
        request.args.get('percentual_divergencia', '')
    )
    fabricante_analisado_para_recontagem_frontend = request.form.get(
        'fabricante_analisado',
        request.args.get('fabricante_analisado', '')
    )
    
    if request.method == 'POST':
        fabricantes_selecionados_for_template = request.form.getlist('fabricantes')
        if 'ver_todos' in request.form:
            fabricantes_selecionados_for_template = list(fabricantes_status_dict.keys())
        if fabricantes_selecionados_for_template:
            inventario_filtrado = [
                item for item in data_manager.INVENTORY_DATA
                if str(item.get('FABRICANTE', '')) in fabricantes_selecionados_for_template
            ]
        if 'analisar_divergencia' in request.form:
            return handle_divergence_analysis(
                fabricantes_selecionados_for_template,
                fabricantes_status_dict,
                percentual_divergencia_str
            )
    
    calculate_item_differences(inventario_filtrado)
    return render_template(
        'index.html',
        inventario=inventario_filtrado,
        fabricantes_status=fabricantes_status_dict,
        fabricantes_selecionados=fabricantes_selecionados_for_template,
        notifications=list(reversed(data_manager.NOTIFICATIONS)),
        percentual_divergencia_aplicado=percentual_divergencia_str,
        fabricante_analisado_para_recontagem=fabricante_analisado_para_recontagem_frontend
    )

def handle_divergence_analysis(fabricantes_selecionados, fabricantes_status_dict, percentual_divergencia_str):
    fab_para_analise = None
    if fabricantes_selecionados:
        for f_sel in fabricantes_selecionados:
            fab_status_obj = fabricantes_status_dict.get(f_sel)
            status_real = fab_status_obj.get('status') if isinstance(fab_status_obj, dict) else fab_status_obj
            if status_real == 'Concluído':
                fab_para_analise = f_sel
                break
        if not fab_para_analise and fabricantes_selecionados:
            fab_para_analise = fabricantes_selecionados[0]
    
    if fab_para_analise:
        return redirect(url_for(
            'gestor.auditoria_recontagem',
            fabricante_analisado=fab_para_analise,
            percentual_referencia=percentual_divergencia_str
        ))
    else:
        flash("Selecione pelo menos um fabricante 'Concluído' para analisar divergências ou use a página de Auditoria.", "warning")
        return redirect(url_for('gestor.index'))

@gestor_bp.route('/auditoria_recontagem', methods=['GET', 'POST'])
@login_required(role="gestor")
@handle_exceptions()
def auditoria_recontagem():
    fabricantes_concluidos_dict = get_completed_manufacturers_with_status()
    produtos_do_fabricante_analisado = []
    fabricante_analisado = request.args.get('fabricante_analisado', request.form.get('fabricante_analisado', ''))
    percentual_referencia_str = request.args.get('percentual_referencia', request.form.get('percentual_referencia', ''))

    # Obtém o nome de exibição com "(verificado)" se disponível
    # O .get() é crucial aqui para pegar o nome de exibição (com "verificado") do dicionário.
    fabricante_analisado_display = fabricantes_concluidos_dict.get(fabricante_analisado, fabricante_analisado)

    if should_analyze_manufacturer(request, fabricante_analisado, percentual_referencia_str):
        produtos_do_fabricante_analisado = process_manufacturer_analysis(
            fabricante_analisado,
            fabricantes_concluidos_dict, # Passa o dicionário completo, se necessário
            percentual_referencia_str
        )
    
    return render_template(
        'auditoria_recontagem.html',
        fabricantes_concluidos=fabricantes_concluidos_dict,
        fabricante_analisado_atual=fabricante_analisado, # Mantém o nome "limpo" para o value do select
        fabricante_analisado_display=fabricante_analisado_display, # Usa este para o título
        percentual_referencia_aplicado=percentual_referencia_str,
        itens_para_analise=produtos_do_fabricante_analisado
    )

def get_completed_manufacturers():
    fabricantes_concluidos_dict = {}
    if not data_manager.INVENTORY_DATA:
        return fabricantes_concluidos_dict

    unique_fabricantes = set(str(item.get('FABRICANTE', '')).strip() for item in data_manager.INVENTORY_DATA if str(item.get('FABRICANTE', '')).strip())
    unique_fabricantes = sorted(unique_fabricantes)

    for fab_nome in unique_fabricantes:
        itens_do_fabricante = [i for i in data_manager.INVENTORY_DATA if str(i.get('FABRICANTE', '')) == fab_nome]
        if itens_do_fabricante and all(str(i.get('STATUS')) == 'Concluído' for i in itens_do_fabricante):
            fabricantes_concluidos_dict[fab_nome] = 'Concluído'

    return fabricantes_concluidos_dict

def get_completed_manufacturers_with_status():
    """
    Retorna fabricantes concluídos com indicação de verificação (se todos os itens estão validados)
    e indicação de recontagem (se houve alguma recontagem).
    A indicação 'verificado' (por validação) tem precedência.
    """
    fabricantes_status_display = {}
    if not data_manager.INVENTORY_DATA:
        return fabricantes_status_display

    # Coleta fabricantes únicos ordenados alfabeticamente
    unique_fabricantes = sorted(set(str(item.get('FABRICANTE', '')).strip() 
                                   for item in data_manager.INVENTORY_DATA 
                                   if str(item.get('FABRICANTE', '')).strip()))

    # Processa cada fabricante
    for fab_nome in unique_fabricantes:
        itens_do_fabricante = [i for i in data_manager.INVENTORY_DATA 
                              if str(i.get('FABRICANTE', '')) == fab_nome]
        
        # Verifica se TODOS os itens do fabricante estão 'Concluído'
        if itens_do_fabricante and all(str(i.get('STATUS')) == 'Concluído' for i in itens_do_fabricante):
            
            # Verifica se TODOS os itens do fabricante estão 'Validado' como 'Sim'
            todos_itens_validados_sim = all(str(item.get('VALIDADO', 'Não')) == 'Sim' for item in itens_do_fabricante)
            
            # Verifica se há **qualquer** item com recontagem (mesmo que sua diferença seja 0 agora)
            tem_recontagem = any(item.get('RECONTAGEM_COUNT', 0) > 0 for item in itens_do_fabricante)
            
            nome_display = fab_nome # Começa com o nome básico

            # Lógica de prioridade: 'verificado' (por validação) tem precedência
            if todos_itens_validados_sim:
                nome_display = f"{nome_display} (verificado)"
            elif tem_recontagem:
                nome_display = f"{nome_display} (com recontagem)"

            fabricantes_status_display[fab_nome] = nome_display
    
    return fabricantes_status_display

def should_analyze_manufacturer(request, fabricante_analisado, percentual_referencia_str):
    return ((request.method == 'POST' and 'analisar_fabricante_btn' in request.form) or
            (request.method == 'GET' and fabricante_analisado and percentual_referencia_str))

def process_manufacturer_analysis(fabricante_analisado, fabricantes_concluidos_dict, percentual_referencia_str):
    produtos_do_fabricante_analisado = []
    
    if not fabricante_analisado:
        flash("Por favor, selecione um fabricante para analisar.", "warning")
        return produtos_do_fabricante_analisado

    if fabricante_analisado not in fabricantes_concluidos_dict:
        flash(f"O fabricante '{fabricante_analisado}' não está totalmente 'Concluído' e não pode ser analisado.", "warning")
        return produtos_do_fabricante_analisado

    for item_original in data_manager.INVENTORY_DATA:
        if (str(item_original.get('FABRICANTE', '')) == fabricante_analisado and
            str(item_original.get('STATUS')) == 'Concluído'):
            item = item_original
            estoque_sistema = item.get('ESTOQUE_SISTEMA', 0)
            inventario_contado = item.get('INVENTARIO', 0)
            diferenca_num = inventario_contado - estoque_sistema
            item['DIFERENCA_NUMERICA_CALC'] = diferenca_num

            # Cálculo de divergência mantido para exibição
            if estoque_sistema != 0:
                divergencia_percentual = (abs(diferenca_num) / float(estoque_sistema)) * 100.0
                item['DIVERGENCIA_PERCENTUAL_CALC'] = round(divergencia_percentual, 2)
            else:
                item['DIVERGENCIA_PERCENTUAL_CALC'] = 'Máxima (Est.Sist. 0)' if diferenca_num != 0 else 0.0

            if 'RECONTAGEM_COUNT' not in item:
                item['RECONTAGEM_COUNT'] = 0

            # ✅ REMOVA OU COMENTE ESTA LINHA:
            # if diferenca_num != 0:  # <--- ESTA É A LINHA DO FILTRO
            produtos_do_fabricante_analisado.append(item) # <-- Esta linha agora será sempre executada para itens concluídos do fabricante.

    if not produtos_do_fabricante_analisado and fabricante_analisado:
        # A mensagem de "Nenhum item com divergência" pode precisar ser ajustada
        # se você espera ver itens "OK" agora.
        flash(f"Nenhum item com divergência encontrado para '{fabricante_analisado}'. Todos os itens estão com estoque correto.", "info")

    return produtos_do_fabricante_analisado

def process_multiple_recount_items(selected_items_raw, fabricante_analisado):
    """Processa múltiplos itens para recontagem. Função auxiliar, não uma rota."""
    current_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    gestor_name = session.get('user_full_name', 'Admin (Sistema)')
    itens_marcados_count = 0
    set_selected_items = set(selected_items_raw)

    for item_data in data_manager.INVENTORY_DATA:
        item_lote_limpo = str(item_data.get('LOTE', '')).strip()
        item_key = f"{item_data.get('CÓDIGO')}|{item_lote_limpo}"
       
        if item_key in set_selected_items:
            # Busca o item fresco do banco/memória para ter valor atualizado
            item_fresco = find_item(item_data.get('CÓDIGO'), item_lote_limpo)
            if not item_fresco:
                continue
            
            # Incrementa o contador de recontagem
            current_count = item_fresco.get('RECONTAGEM_COUNT', 0)
            new_count = current_count + 1
            
            contador_original_val = item_fresco.get('CONTADOR_ORIGINAL', '') or item_fresco.get('CONTADOR', 'Sistema')
            
            # Atualiza CONFERENCIA com histórico usando o item fresco
            conferencia_atual = item_fresco.get('CONFERENCIA', '')
            nova_conferencia = f"{conferencia_atual} | Recontagem {new_count}: solicitada por {gestor_name} em {current_time}".strip('| ')
            
            updates = {
                'STATUS': 'Recontagem',
                'CONTADOR_ORIGINAL': contador_original_val,
                'HORARIO_RECONTAGEM_SOLICITADA': current_time,
                'GESTOR_RECONTAGEM': gestor_name,
                'CONFERENCIA': nova_conferencia,
                'ESTOQUE_BALCAO': 0,
                'ESTOQUE_DEPOSITO': 0,
                'INVENTARIO': 0,
                'CONTADOR': '',
                'RECONTAGEM_COUNT': new_count  # Salva o contador incrementado
            }
            
            update_item(item_data.get('CÓDIGO'), item_lote_limpo, updates)
            itens_marcados_count += 1

    if itens_marcados_count > 0:
        data_manager.NOTIFICATIONS.append({
            'user': f"Sistema (Gestor: {gestor_name})",
            'manufacturer': f"{fabricante_analisado}",
            'message': f"Recontagem de {itens_marcados_count} item(s) de '{fabricante_analisado}' solicitada.",
            'timestamp': current_time
        })
   
    return itens_marcados_count

# --- NOVA FUNÇÃO E ROTA UNIFICADA PARA TODAS AS AÇÕES EM LOTE ---
@gestor_bp.route('/processar_acoes_em_lote', methods=['POST'])
@login_required(role="gestor")
@handle_exceptions()
def processar_acoes_em_lote():
    """Processa múltiplas ações (recontagem e validação/invalidação) em lote."""
    selected_for_recount_raw = request.form.getlist('selected_for_recount')
    selected_for_validate_raw = request.form.getlist('selected_for_validate')
    # No seu HTML, 'selected_for_invalidate' não é enviado explicitamente como um array separado de checkboxes.
    # A lógica é que o checkbox 'item-checkbox-validate' marca para 'Sim'.
    # Para lidar com 'Não', a lógica deveria ser mais complexa (saber o estado anterior do item).
    # Se você quiser "Invalidar" explicitamente, precisaria de um checkbox ou botão separado no HTML
    # ou uma lógica no JS que detecte "desmarcado de Sim para Não".
    # Por ora, vamos assumir que selected_for_validate_raw é para marcar 'Sim'.
    # E vamos ignorar selected_for_invalidate_raw, pois o HTML não o preenche.
    # Se precisar de invalidar, teremos que rever o frontend.
    selected_for_invalidate_raw = [] # Garante que está vazio se não vier do frontend

    fabricante_analisado_hidden = request.form.get('fabricante_analisado_hidden_form', '')
    percentual_divergencia_hidden = request.form.get('percentual_divergencia_hidden_form', '')

    total_acoes_realizadas = 0
    messages = []

    # Processar itens para Recontagem
    if selected_for_recount_raw:
        recounted_count = process_multiple_recount_items(selected_for_recount_raw, fabricante_analisado_hidden)
        if recounted_count > 0:
            messages.append(f"{recounted_count} item(s) enviado(s) para recontagem.")
            total_acoes_realizadas += recounted_count

    # Processar itens para Validação (marcar como 'Sim')
    if selected_for_validate_raw:
        validated_count = process_multiple_validation_items(selected_for_validate_raw, 'Sim')
        if validated_count > 0:
            messages.append(f"{validated_count} item(s) marcado(s) como 'Validado'.")
            total_acoes_realizadas += validated_count

    # Processar itens para Invalidação (marcar como 'Não')
    # Esta parte só seria acionada se 'selected_for_invalidate_raw' viesse preenchido do frontend
    # com itens que deveriam ser explicitamente marcados como 'Não'.
    if selected_for_invalidate_raw: # Apenas se o frontend realmente enviar isso
        invalidated_count = process_multiple_validation_items(selected_for_invalidate_raw, 'Não')
        if invalidated_count > 0:
            messages.append(f"{invalidated_count} item(s) marcado(s) como 'Não Validado'.")
            total_acoes_realizadas += invalidated_count

    if total_acoes_realizadas > 0:
        flash("Ações em lote concluídas: " + " ".join(messages), "success")
    else:
        flash("Nenhuma ação em lote foi processada. Nenhum item válido selecionado para recontagem ou validação.", "info")

    # --- É ESSENCIAL RECARREGAR OS DADOS APÓS AS ALTERAÇÕES NO DB ---
    data_manager.load_or_create_inventory() # Recarrega os dados em memória do DB para refletir as mudanças

    return redirect(url_for('gestor.auditoria_recontagem',
                           fabricante_analisado=fabricante_analisado_hidden,
                           percentual_referencia=percentual_divergencia_hidden))

# --- Nova Função Auxiliar para Processar Múltiplos Itens de Validação ---
def process_multiple_validation_items(selected_items_raw, validation_status):
    """Processa múltiplos itens para marcar como validado/não validado. Função auxiliar, não uma rota."""
    current_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    gestor_name = session.get('user_full_name', 'Admin (Sistema)')
    items_processed_count = 0
    set_selected_items = set(selected_items_raw)

    for item_data in data_manager.INVENTORY_DATA:
        item_lote_limpo = str(item_data.get('LOTE', '')).strip()
        item_key = f"{item_data.get('CÓDIGO')}|{item_lote_limpo}"

        if item_key in set_selected_items:
            item_fresco = find_item(item_data.get('CÓDIGO'), item_lote_limpo)
            if not item_fresco:
                continue

            # Atualiza o status 'VALIDADO'
            updates = {'VALIDADO': validation_status}
            if update_item(item_data.get('CÓDIGO'), item_lote_limpo, updates):
                items_processed_count += 1
                data_manager.NOTIFICATIONS.append({
                    'user': f"Sistema (Gestor: {gestor_name})",
                    'manufacturer': f"{item_fresco.get('FABRICANTE')}",
                    'message': f"Item {item_fresco.get('PRODUTO')} (Cód: {item_data.get('CÓDIGO')}, Lote: {item_lote_limpo or 'N/A'}) validado para '{validation_status}'.",
                    'timestamp': current_time
                })
    return items_processed_count

# REMOVIDA: @gestor_bp.route('/marcar_multiplos_invalidado', methods=['POST'])
# REMOVIDA: def marcar_multiplos_invalidado(): ...

@gestor_bp.route('/atualizar_planilha_existente', methods=['POST'])
@login_required(role="gestor")
@handle_exceptions()
def atualizar_planilha_existente():
    file = request.files.get('planilha_excel')
    if not file or file.filename == '':
        flash('Nenhum arquivo selecionado.', 'warning')
        return redirect(url_for('gestor.index'))
    
    if not file.filename.endswith('.xlsx'):
        flash('Tipo de arquivo inválido. Envie um arquivo .xlsx', 'warning')
        return redirect(url_for('gestor.index'))
    
    workbook = openpyxl.load_workbook(file)
    sheet_name = EXCEL_CONFIG['NOME_ABA_EXCEL']
    
    if sheet_name not in workbook.sheetnames:
        flash(f"Aba '{sheet_name}' não encontrada.", "danger")
        return redirect(url_for('gestor.index'))
    
    sheet = workbook[sheet_name]
    num_updated, num_added = process_excel_optimized(sheet)
    total_alteracoes = num_updated + num_added
    
    if total_alteracoes == 0:
        flash("Nenhuma linha foi atualizada ou adicionada.", "info")
    else:
        flash_message_parts = []
        if num_updated > 0:
            flash_message_parts.append(f"{num_updated} linha(s) existente(s) atualizada(s)")
        if num_added > 0:
            flash_message_parts.append(f"{num_added} nova(s) linha(s) inserida(s)")
        flash(f"Sucesso! { ' e '.join(flash_message_parts) }.", "success")
    
    return generate_excel_download(workbook, file.filename)

def process_excel_optimized(sheet):
    column_indices, headers = get_column_indices(sheet)
    if not column_indices or not headers:
        return 0, 0
    
    inventory_index = create_inventory_index()
    if not inventory_index:
        return 0, 0
    
    updated_keys = process_excel_rows_batch(sheet, column_indices, inventory_index)
    novas_linhas_inseridas = insert_new_items_into_sheet(sheet, inventory_index, updated_keys, headers, column_indices)
    
    return len(updated_keys), novas_linhas_inseridas

def get_column_indices(sheet):
    try:
        headers = [str(cell.value if cell.value is not None else '').strip() for cell in sheet[EXCEL_CONFIG['LINHA_CABECALHO_EXCEL']]]
        required_columns = {
            'codigo': EXCEL_CONFIG['COL_EXCEL_CHAVE_CODIGO'], 
            'lote': EXCEL_CONFIG['COL_EXCEL_CHAVE_LOTE'],
            'resp_contagem': EXCEL_CONFIG['COL_EXCEL_DESTINO_RESP_CONTAGEM'], 
            'inventario': EXCEL_CONFIG['COL_EXCEL_DESTINO_INVENTARIO'],
            'conferencia': EXCEL_CONFIG['COL_EXCEL_DESTINO_CONFERENCIA']
        }
        
        column_indices = {}
        for key, col_name in required_columns.items():
            try:
                column_indices[key] = headers.index(col_name) + 1
            except ValueError:
                flash(f"Coluna obrigatória '{col_name}' não encontrada na planilha.", "danger")
                return None, None
        
        return column_indices, headers
    except Exception as e:
        logger.error(f"Erro ao processar cabeçalhos do Excel: {e}")
        return None, None

def create_inventory_index():
    inventory_index = {}
    for item in data_manager.INVENTORY_DATA:
        if item.get('STATUS') == 'Concluído':
            codigo = str(item.get('CÓDIGO', '')).replace('.0', '').zfill(7).strip()
            lote = str(item.get('LOTE', '')).strip() if item.get('LOTE') and str(item.get('LOTE')).lower() != 'nan' else ''
            if codigo:
                key = f"{codigo}|{lote}"
                inventory_index[key] = item
    return inventory_index

def process_excel_rows_batch(sheet, column_indices, inventory_index):
    updated_keys = set()
    start_row = EXCEL_CONFIG['LINHA_CABECALHO_EXCEL'] + 1
    
    for row_idx in range(start_row, sheet.max_row + 1):
        try:
            codigo_val = sheet.cell(row=row_idx, column=column_indices['codigo']).value
            lote_val = sheet.cell(row=row_idx, column=column_indices['lote']).value
            
            if codigo_val is not None:
                codigo = str(codigo_val).replace('.0', '').zfill(7).strip()
                lote = str(lote_val).strip() if lote_val and str(lote_val).lower() != 'nan' else ''
                key = f"{codigo}|{lote}"
                
                if key in inventory_index:
                    item_sistema = inventory_index[key]
                    if update_excel_row_optimized(sheet, row_idx, column_indices, item_sistema):
                        updated_keys.add(key)
        except Exception as e:
            logger.warning(f"Erro ao processar linha {row_idx} da planilha: {e}")
            continue
    
    return updated_keys

def update_excel_row_optimized(sheet, row_idx, column_indices, item_sistema):
    try:
        resp_contagem = str(item_sistema.get('CONTADOR', ''))
        inventario_val = item_sistema.get('INVENTARIO')
        conferencia_info = build_conference_info_optimized(item_sistema)
        
        update_cell_preserve_format(sheet, row_idx, column_indices['resp_contagem'], resp_contagem)
        update_cell_preserve_format(sheet, row_idx, column_indices['inventario'], inventario_val if inventario_val is not None else '')
        update_cell_preserve_format(sheet, row_idx, column_indices['conferencia'], conferencia_info)
        
        return True
    except Exception as e:
        logger.warning(f"Falha ao atualizar a linha {row_idx}: {e}")
        return False

def update_cell_preserve_format(sheet, row, col, value):
    cell = sheet.cell(row=row, column=col)
    if cell.has_style:
        original_style = cell._style
        cell.value = value
        cell._style = original_style
    else:
        cell.value = value

def build_conference_info_optimized(item_sistema):
    parts = []
    gestor = item_sistema.get('GESTOR_RECONTAGEM', '')
    horario = item_sistema.get('HORARIO_RECONTAGEM_SOLICITADA', '')
    contador_original = item_sistema.get('CONTADOR_ORIGINAL', '')
    obs_conferencia_original = item_sistema.get('CONFERENCIA', '') 
    
    if gestor and horario and str(gestor).lower() not in ['nan', 'none', ''] and str(horario).lower() not in ['nan', 'none', '']:
        recount_info = f"Recontagem solicitada por {gestor} em {horario}"
        if contador_original and str(contador_original).lower() not in ['nan', 'none', '']:
            recount_info += f" (Contador Original: {contador_original})"
        parts.append(recount_info)

    if obs_conferencia_original and \
       str(obs_conferencia_original).lower() not in ['nan', 'none', '']:
        
        if not ("Recontagem solicitada por" in obs_conferencia_original or
                "Recontagem (múltipla) por" in obs_conferencia_original) or \
           "Lote adicionado por" in obs_conferencia_original:
            parts.append(obs_conferencia_original)

    return " - ".join(parts)

def insert_new_items_into_sheet(sheet, inventory_index, updated_keys, headers, column_indices):
    items_to_add_keys = set(inventory_index.keys()) - updated_keys
    if not items_to_add_keys:
        return 0
    
    items_to_process = []
    start_row = EXCEL_CONFIG['LINHA_CABECALHO_EXCEL']
    
    for key in items_to_add_keys:
        item_data = inventory_index[key]
        item_codigo = str(item_data.get('CÓDIGO', '')).replace('.0', '').zfill(7).strip()
        insertion_row = -1
        
        for row_idx in range(sheet.max_row, start_row, -1):
            cell_val = sheet.cell(row=row_idx, column=column_indices['codigo']).value
            codigo_in_sheet = str(cell_val).replace('.0', '').zfill(7).strip() if cell_val else ''
            if item_codigo == codigo_in_sheet:
                insertion_row = row_idx + 1
                break
        
        items_to_process.append({'key': key, 'insertion_row': insertion_row, 'data': item_data})
    
    items_to_process.sort(key=lambda x: x['insertion_row'], reverse=True)
    count_added = 0
    
    for item_job in items_to_process:
        insertion_row = item_job['insertion_row']
        item_data = item_job['data']
        
        data_to_overwrite = {
            'DATA': datetime.now().strftime("%d/%m/%Y"),
            'RESP. PELA CONTAGEM': item_data.get('CONTADOR', ''),
            'LOTE': item_data.get('LOTE', ''),
            'INVENTÁRIO': item_data.get('INVENTARIO', 0),
            'CONFERENCIA': build_conference_info_optimized(item_data),
            'FAB E VAL': item_data.get('DATA_VALIDADE', '')
        }
        
        if insertion_row != -1:
            sheet.insert_rows(insertion_row)
            target_row_to_write = insertion_row
            source_row_to_copy_style = insertion_row - 1
        else:
            sheet.append([''] * len(headers))
            target_row_to_write = sheet.max_row
            source_row_to_copy_style = sheet.max_row - 1
        
        for col_idx, header in enumerate(headers, 1):
            dest_cell = sheet.cell(row=target_row_to_write, column=col_idx)
            source_cell = sheet.cell(row=source_row_to_copy_style, column=col_idx)
            
            if source_cell.has_style:
                dest_cell.font = copy(source_cell.font)
                dest_cell.border = copy(source_cell.border)
                dest_cell.fill = copy(source_cell.fill)
                dest_cell.number_format = copy(source_cell.number_format)
                dest_cell.protection = copy(source_cell.protection)
                dest_cell.alignment = copy(source_cell.alignment)
            
            if header in data_to_overwrite:
                dest_cell.value = data_to_overwrite[header]
            elif header not in ['DIFERENÇA', 'POSIÇÃO']:
                dest_cell.value = source_cell.value
        
        count_added += 1
    
    return count_added

def generate_excel_download(workbook, original_filename):
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    safe_filename = secure_filename(original_filename)
    download_filename = f"atualizada_{safe_filename}"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=download_filename)

@gestor_bp.route('/download/planilha_preenchida_recontagem')
@login_required(role="gestor")
@handle_exceptions()
def download_planilha_preenchida_recontagem():
    if not data_manager.INVENTORY_DATA:
        flash("Nenhum dado de inventário disponível.", "warning")
        return redirect(url_for('gestor.index'))
    
    dados_para_planilha = prepare_recount_data()
    if not dados_para_planilha:
        flash("Nenhum item aplicável encontrado para a planilha.", "info")
        return redirect(url_for('gestor.index'))
    
    return create_excel_download(dados_para_planilha, 'DadosConferencia', 'inventario_dados_conferencia.xlsx')

def prepare_recount_data():
    dados_para_planilha = []
    for item in data_manager.INVENTORY_DATA:
        if item.get('STATUS') == 'Concluído' or item.get('CONTADOR_ORIGINAL'):
            conferencia_info = build_conference_info(item)
            dados_para_planilha.append({
                'CÓDIGO': item.get('CÓDIGO'),
                'LOTE': item.get('LOTE', ''),
                'PRODUTO': item.get('PRODUTO'),
                'RESP. PELA CONTAGEM': item.get('CONTADOR'),
                'INVENTÁRIO': item.get('INVENTARIO'),
                'CONFERENCIA': conferencia_info,
                'FABRICANTE': item.get('FABRICANTE'),
                'STATUS ATUAL': item.get('STATUS'),
                'CONTADOR ORIGINAL': item.get('CONTADOR_ORIGINAL', '')
            })
    return dados_para_planilha

def build_conference_info(item):
    conferencia_info = ''
    gestor_recontagem = item.get('GESTOR_RECONTAGEM', '')
    horario_recontagem = item.get('HORARIO_RECONTAGEM_SOLICITADA', '')
    contador_original = item.get('CONTADOR_ORIGINAL', '')
    contador_atual = item.get('CONTADOR', '')
    obs_conferencia = str(item.get('CONFERENCIA', '')).strip()

    gestor_recontagem = str(gestor_recontagem).strip() if gestor_recontagem and str(gestor_recontagem).lower() not in ['nan', 'none', ''] else ''
    horario_recontagem = str(horario_recontagem).strip() if horario_recontagem and str(horario_recontagem).lower() not in ['nan', 'none', ''] else ''
    contador_original = str(contador_original).strip() if contador_original and str(contador_original).lower() not in ['nan', 'none', ''] else ''
    contador_atual = str(contador_atual).strip() if contador_atual and str(contador_atual).lower() not in ['nan', 'none', ''] else ''


    parts = []
    if gestor_recontagem and horario_recontagem:
        recount_details = f"Recontagem solicitada por {gestor_recontagem} em {horario_recontagem}"
        if contador_original:
            recount_details += f" (Contador Original: {contador_original})"
        parts.append(recount_details)

    if obs_conferencia and obs_conferencia.lower() not in ['nan', 'none', '']:
        if not ("Recontagem solicitada por" in obs_conferencia or
                "Recontagem (múltipla) por" in obs_conferencia) or \
           "Lote adicionado por" in obs_conferencia:
            parts.append(obs_conferencia)
            
    if contador_atual and contador_atual != contador_original and "Lote adicionado por" not in obs_conferencia:
        if parts:
            parts.append(f"Contado por: {contador_atual}")
        else:
            parts.append(f"Contado por: {contador_atual}")

    return " - ".join(parts)

@gestor_bp.route('/download/ajustes')
@login_required(role="gestor")
@handle_exceptions()
def download_ajustes():
    if not data_manager.INVENTORY_DATA:
        flash("Nenhum dado de inventário.", "warning")
        return redirect(url_for('gestor.index'))
    
    df_export_final = prepare_adjustments_data()
    if df_export_final.empty:
        flash("Nenhum item 'Concluído' para gerar o relatório de ajustes.", "info")
        return redirect(url_for('gestor.index'))
    
    return create_dataframe_download(df_export_final, 'AjustesAgregados', 'relatorio_ajustes_agregados.xlsx')

def prepare_adjustments_data():
    df_concluidos = pd.DataFrame([
        item for item in data_manager.INVENTORY_DATA if item.get('STATUS') == 'Concluído'
    ])
    
    if df_concluidos.empty:
        return df_concluidos
    
    df_concluidos['ESTOQUE_SISTEMA'] = pd.to_numeric(df_concluidos['ESTOQUE_SISTEMA'], errors='coerce').fillna(0)
    df_concluidos['INVENTARIO'] = pd.to_numeric(df_concluidos['INVENTARIO'], errors='coerce').fillna(0)
    
    df_agrupado = df_concluidos.groupby(
        ['CÓDIGO', 'PRODUTO', 'FABRICANTE'], as_index=False
    ).agg(
        QUANTIDADE_SISTEMA=('ESTOQUE_SISTEMA', 'sum'),
        INVENTARIO_TOTAL=('INVENTARIO', 'sum')
    )
    
    df_agrupado['DIFERENCA_AJUSTE'] = df_agrupado['INVENTARIO_TOTAL'] - df_agrupado['QUANTIDADE_SISTEMA']
    df_agrupado['POSICAO_AJUSTE'] = df_agrupado['DIFERENCA_AJUSTE'].apply(
        lambda x: 'Mais' if x > 0 else ('Menos' if x < 0 else 'OK')
    )
    df_agrupado['DATA_RELATORIO'] = datetime.now().strftime("%d/%m/%Y")
    
    df_agrupado.rename(columns={
        'FABRICANTE': 'FORNECEDOR',
        'PRODUTO': 'DESCRIÇÃO DO PRODUTO/SERVIÇO',
        'QUANTIDADE_SISTEMA': 'QTD. SISTEMA',
        'INVENTARIO_TOTAL': 'QTD. INVENTARIADA',
        'DIFERENCA_AJUSTE': 'DIFERENÇA',
        'POSICAO_AJUSTE': 'POSIÇÃO',
        'DATA_RELATORIO': 'DATA'
    }, inplace=True)
    
    colunas_export = ['DATA', 'FORNECEDOR', 'CÓDIGO', 'DESCRIÇÃO DO PRODUTO/SERVIÇO',
                     'QTD. SISTEMA', 'QTD. INVENTARIADA', 'DIFERENÇA', 'POSIÇÃO']
    
    return df_agrupado[[col for col in colunas_export if col in df_agrupado.columns]]

@gestor_bp.route('/download/selecionados')
@login_required(role="gestor")
@handle_exceptions()
def download_selecionados():
    fabricantes_selecionados_req = request.args.getlist('fab')
    if not fabricantes_selecionados_req:
        flash("Nenhum fabricante selecionado para download.", "warning")
        return redirect(url_for('gestor.index'))
    
    itens_selecionados = [
        item for item in data_manager.INVENTORY_DATA
        if str(item.get('FABRICANTE', '')) in fabricantes_selecionados_req
    ]
    
    if not itens_selecionados:
        flash("Nenhum item encontrado para os fabricantes selecionados.", "info")
        return redirect(url_for('gestor.index'))
    
    df_export_final = prepare_selected_manufacturers_data(itens_selecionados)
    return create_dataframe_download(df_export_final, 'InventarioSelecionado', 'inventario_fabricantes_selecionados.xlsx')

def prepare_selected_manufacturers_data(itens_selecionados):
    df_export = pd.DataFrame(itens_selecionados)
    cols_display = ['CÓDIGO', 'PRODUTO', 'LOTE', 'FABRICANTE', 'ESTOQUE_SISTEMA', 'INVENTARIO',
                   'STATUS', 'CONTADOR', 'CONTADOR_ORIGINAL', 'HORARIO_RECONTAGEM_SOLICITADA',
                   'GESTOR_RECONTAGEM', 'CONFERENCIA']
    
    df_export_final = df_export[[col for col in cols_display if col in df_export.columns]].copy()
    
    if 'INVENTARIO' in df_export_final.columns and 'ESTOQUE_SISTEMA' in df_export_final.columns:
        df_export_final['DIFERENCA'] = (
            pd.to_numeric(df_export_final['INVENTARIO'], errors='coerce').fillna(0) -
            pd.to_numeric(df_export_final['ESTOQUE_SISTEMA'], errors='coerce').fillna(0)
        )
    else:
        df_export_final['DIFERENCA'] = 0
    
    return df_export_final

@gestor_bp.route('/download/novos_lotes')
@login_required("gestor")
@handle_exceptions()
def download_novos_lotes():
    """Gera relatório Excel com itens que possuem DATA DE FABRICAÇÃO e DATA DE VALIDADE preenchidas."""
    try:
        # Filtra itens que possuem ambas as datas preenchidas
        itens_novos_lotes = []
        
        for item in data_manager.INVENTORY_DATA:
            data_fabricacao = item.get('DATA_FABRICACAO', '')
            data_validade = item.get('DATA_VALIDADE', '')
            
            # Converte para string e remove espaços
            fab_str = str(data_fabricacao).strip().upper() if data_fabricacao else ''
            val_str = str(data_validade).strip().upper() if data_validade else ''
            
            # Lista de valores considerados vazios
            valores_vazios = ['', 'N/A', 'NULL', 'NONE', 'NAN', '-', '0']
            
            # Verifica se ambas as datas estão preenchidas
            fab_preenchida = fab_str not in valores_vazios and len(fab_str) > 0
            val_preenchida = val_str not in valores_vazios and len(val_str) > 0
            
            if fab_preenchida and val_preenchida:
                itens_novos_lotes.append(item)
        
        if not itens_novos_lotes:
            flash('Nenhum item com datas de fabricação e validade encontrado.', 'warning')
            return redirect(url_for('gestor.index'))
        
        # Cria DataFrame com os itens filtrados
        df = pd.DataFrame(itens_novos_lotes)
        
        # Seleciona e ordena as colunas para o relatório
        colunas_relatorio = [
            'CÓDIGO', 'PRODUTO', 'FABRICANTE', 'LOTE',
            'ESTOQUE_SISTEMA', 'INVENTARIO', 'DATA_FABRICACAO', 
            'DATA_VALIDADE', 'CONTADOR', 'STATUS',
            'HORARIO_RECONTAGEM_SOLICITADA', 'GESTOR_RECONTAGEM'
        ]
        
        # Filtra apenas as colunas que existem no DataFrame
        colunas_existentes = [col for col in colunas_relatorio if col in df.columns]
        df_relatorio = df[colunas_existentes].copy()
        
        # Ordena por fabricante e depois por código
        df_relatorio = df_relatorio.sort_values(['FABRICANTE', 'CÓDIGO'])
        
        # Formata as datas se necessário
        for col_data in ['DATA_FABRICACAO', 'DATA_VALIDADE']:
            if col_data in df_relatorio.columns:
                df_relatorio[col_data] = df_relatorio[col_data].astype(str)
        
        # Cria arquivo Excel na memória
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Escreve os dados
            df_relatorio.to_excel(
                writer, 
                sheet_name='Novos Lotes', 
                index=False,
                startrow=3  # Deixa espaço para cabeçalho
            )
            
            # Acessa a planilha para formatação
            worksheet = writer.sheets['Novos Lotes']
            
            # Adiciona cabeçalho do relatório
            worksheet['A1'] = 'RELATÓRIO DE NOVOS LOTES'
            worksheet['A2'] = f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M:%S")}'
            worksheet['A3'] = f'Total de itens: {len(df_relatorio)}'
            
            # Título principal
            worksheet['A1'].font = Font(bold=True, size=16)
            worksheet['A1'].alignment = Alignment(horizontal='left')
            
            # Informações
            worksheet['A2'].font = Font(size=12)
            worksheet['A3'].font = Font(size=12, bold=True)
            
            # Formata cabeçalhos das colunas (linha 4)
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            for col_num, column in enumerate(df_relatorio.columns, 1):
                cell = worksheet.cell(row=4, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-ajusta largura das colunas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Nome do arquivo com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"relatorio_novos_lotes_{timestamp}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Erro ao gerar relatório de novos lotes: {e}")
        flash(f'Erro ao gerar relatório de novos lotes: {str(e)}', 'danger')
        return redirect(url_for('gestor.index'))

def create_excel_download(data, sheet_name, filename):
    df_export = pd.DataFrame(data)
    return create_dataframe_download(df_export, sheet_name, filename)

def create_dataframe_download(df, sheet_name, filename):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@gestor_bp.route('/get_updates')
@login_required(role="gestor")
def get_updates():
    num_seen_notifications = session.get('seen_notifications', 0)
    new_notification_count = len(data_manager.NOTIFICATIONS) - num_seen_notifications
    current_fabricantes_status = get_manufacturers_status()
    chart_data = generate_chart_data(current_fabricantes_status)
    
    return jsonify({
        'manufacturers_status': current_fabricantes_status,
        'notifications': list(reversed(data_manager.NOTIFICATIONS)),
        'new_notification_count': new_notification_count,
        'chart1_progresso': chart_data['progress'],
        'chart2_diferencas': chart_data['differences']
    })

def generate_chart_data(fabricantes_status):
    chart1_labels = ['Concluído', 'Em Andamento/Pendente']
    chart1_data = [0, 0]
    
    if fabricantes_status:
        for fab_info in fabricantes_status.values():
            if fab_info['status'] == 'Concluído':
                chart1_data[0] += 1
            else:
                chart1_data[1] += 1
    
    chart2_labels = ['OK', 'Menos', 'Mais']
    chart2_data = [0, 0, 0]
    
    if data_manager.INVENTORY_DATA:
        for item in data_manager.INVENTORY_DATA:
            if str(item.get('STATUS')) == 'Concluído':
                inventario_val = item.get('INVENTARIO', 0)
                estoque_sistema_val = item.get('ESTOQUE_SISTEMA', 0)
                diferenca = inventario_val - estoque_sistema_val
                if diferenca == 0:
                    chart2_data[0] += 1
                elif diferenca < 0:
                    chart2_data[1] += 1
                else:
                    chart2_data[2] += 1
    
    return {
        'progress': {'labels': chart1_labels, 'data': chart1_data},
        'differences': {'labels': chart2_labels, 'data': chart2_data}
    }

@gestor_bp.route('/mark_notifications_seen', methods=['POST'])
@login_required(role="gestor")
def mark_notifications_seen():
    mark_notifications_as_seen()
    return jsonify(success=True)

@gestor_bp.route('/api/estatisticas')
@login_required(role="gestor")
def api_estatisticas():
    """API para retornar estatísticas atualizadas em JSON."""
    try:
        # Recalcula as estatísticas dos fabricantes
        fabricantes_status = get_manufacturers_status()  # Usa sua função existente

        total = len(fabricantes_status)
        pendentes = sum(1 for fab_info in fabricantes_status.values() if fab_info.get('status') == 'Pendente')
        em_andamento = sum(1 for fab_info in fabricantes_status.values() if fab_info.get('status') == 'Em Andamento') # Add this line
        concluidos = sum(1 for fab_info in fabricantes_status.values() if fab_info.get('status') == 'Concluído')
        verificados = sum(1 for fab_info in fabricantes_status.values() if fab_info.get('status_verificado', False))


        # Calcula estatísticas de inventário
        inventario_stats = {
            'total_itens': len(data_manager.INVENTORY_DATA),
            'itens_concluidos': len([item for item in data_manager.INVENTORY_DATA if item.get('STATUS') == 'Concluído']),
            'notificacoes': len(data_manager.NOTIFICATIONS)
        }

        return jsonify({
            'success': True,
            'timestamp': datetime.now().strftime('%H:%M:%S'),
            'fabricantes': {
                'total': total,
                'pendentes': pendentes,
                'em_andamento': em_andamento, # Add this line
                'concluidos': concluidos,
                'verificados': verificados # Adicione esta linha
            },
            'inventario': inventario_stats
        })

    except Exception as e:
        logger.error(f"Erro na API de estatísticas: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'timestamp': datetime.now().strftime('%H:%M:%S')
        }), 500
@gestor_bp.route('/redefinir_senha', methods=['GET', 'POST'])
@login_required(role="gestor")
@handle_exceptions()
def redefinir_senha():
    if request.method == 'POST':
        user_id = request.form.get('user_id')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')

        if not all([user_id, new_password, confirm_password]):
            flash('Todos os campos são obrigatórios.', 'warning')
            return redirect(url_for('gestor.redefinir_senha'))

        if new_password != confirm_password:
            flash('As senhas não coincidem.', 'danger')
            return redirect(url_for('gestor.redefinir_senha'))
        
        if len(new_password) < 6:
            flash('A senha deve ter no mínimo 6 caracteres.', 'warning')
            return redirect(url_for('gestor.redefinir_senha'))

        # Gera o hash da nova senha
        password_hash = generate_password_hash(new_password, method='pbkdf2:sha256')

        # Atualiza a senha no banco de dados
        if update_user_password(user_id, password_hash):
            flash('Senha redefinida com sucesso!', 'success')
        else:
            flash('Ocorreu um erro ao redefinir a senha.', 'danger')
        
        return redirect(url_for('gestor.redefinir_senha'))

    # Método GET: Carrega a página
    users = get_all_users()
    # Remove o próprio gestor logado da lista para que ele não redefina a própria senha nesta tela
    current_user_id = session.get('user_id')
    users_filtered = [user for user in users if user['id'] != current_user_id]
    
    return render_template('redefinir_senha.html', users=users_filtered)